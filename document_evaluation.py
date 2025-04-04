#!/usr/bin/env python3
"""
Script to generate an evaluation Excel spreadsheet for the Document Research Agent.
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Document types and their associated files
DOCUMENT_TYPES = {
    "Employment Contracts": [
        "test_data/legal_document_01_employment_contract.pdf",
        "test_data/legal_document_06_employment_contract.pdf"
    ],
    "Service Agreements": [
        "test_data/legal_document_02_service_agreement.pdf", 
        "test_data/legal_document_07_service_agreement.pdf"
    ],
    "Non-Disclosure Agreements": [
        "test_data/legal_document_03_non_disclosure_agreement.pdf",
        "test_data/legal_document_08_non_disclosure_agreement.pdf",
        "test_data/sample_nda.pdf"
    ],
    "Lease Agreements": [
        "test_data/legal_document_04_lease_agreement.pdf",
        "test_data/legal_document_09_lease_agreement.pdf"
    ],
    "Purchase Agreements": [
        "test_data/legal_document_05_purchase_agreement.pdf",
        "test_data/legal_document_10_purchase_agreement.pdf"
    ],
    "Shuttle Service Contract": [
        "test_data/sample_contract_shuttle.pdf"
    ]
}

# Define 5 questions for each document type
EVALUATION_QUESTIONS = {
    "Employment Contracts": [
        {
            "question": "What is the duration of the employment term?",
            "expected_answer": "The employment term varies by contract but is typically specified as a period of years (e.g., 3 years, 5 years) commencing on a specific start date."
        },
        {
            "question": "What confidentiality obligations does the employee have?",
            "expected_answer": "The employee acknowledges access to trade secrets and confidential information owned by the employer and is obligated to maintain confidentiality."
        },
        {
            "question": "What compensation is provided to the employee?",
            "expected_answer": "The employee receives a base salary (typically between $85,000-$120,000 annually) payable according to the employer's normal payroll practices."
        },
        {
            "question": "What benefits is the employee entitled to?",
            "expected_answer": "The employee is eligible to participate in all employee benefit plans, programs, and arrangements generally made available to other employees of the employer."
        },
        {
            "question": "Which state's laws govern the employment agreement?",
            "expected_answer": "The agreement is governed by the laws of a specific state (e.g., California, New York, Texas, Florida, or Illinois) as specified in the governing law section."
        }
    ],
    "Service Agreements": [
        {
            "question": "What services does the service provider agree to perform?",
            "expected_answer": "The service provider agrees to provide specific services such as IT consulting, marketing services, legal services, accounting services, or design services as specified in the agreement."
        },
        {
            "question": "What are the payment terms for the services?",
            "expected_answer": "Payment terms vary by agreement but typically include a fee amount (e.g., $5,000-$15,000) with payment scheduled as monthly installments, upon completion, 50% upfront with 50% upon completion, quarterly installments, or weekly installments."
        },
        {
            "question": "What is the relationship between the service provider and the client?",
            "expected_answer": "The service provider is an independent contractor and not an employee of the client. The service provider is responsible for taxes, insurance, and other obligations related to the services provided."
        },
        {
            "question": "How can the agreement be terminated?",
            "expected_answer": "Either party may terminate the agreement upon a specified notice period (typically 15-90 days) in writing to the other party. Upon termination, the client pays for services performed up to the termination date."
        },
        {
            "question": "What are the confidentiality obligations of the service provider?",
            "expected_answer": "During the term of the agreement and thereafter, the service provider must maintain the confidentiality of any proprietary or confidential information of the client."
        }
    ],
    "Non-Disclosure Agreements": [
        {
            "question": "What is the definition of confidential information in the agreement?",
            "expected_answer": "Confidential Information means information disclosed by the Disclosing Party to the Receiving Party, directly or indirectly, in writing, orally, or by inspection of tangible objects, that is designated as 'Confidential', 'Proprietary' or similarly designated."
        },
        {
            "question": "What are the non-disclosure obligations of the receiving party?",
            "expected_answer": "The Receiving Party must hold Confidential Information in strict confidence and not disclose such Confidential Information to any third party."
        },
        {
            "question": "How long do the confidentiality obligations last?",
            "expected_answer": "The obligations survive termination of any business relationship between the parties and continue for a specified period (typically several years) from the date of disclosure."
        },
        {
            "question": "What happens to confidential materials at the end of the agreement?",
            "expected_answer": "All documents and tangible objects containing Confidential Information remain the property of the Disclosing Party and must be promptly returned upon written request."
        },
        {
            "question": "Does the agreement grant any intellectual property rights?",
            "expected_answer": "No. The agreement explicitly states that nothing shall be construed as granting any rights to the Receiving Party under any patent, copyright, or other intellectual property right of the Disclosing Party."
        }
    ],
    "Lease Agreements": [
        {
            "question": "What is the monthly rent amount for the leased premises?",
            "expected_answer": "The monthly rent varies by agreement but typically ranges from $4,000 to $8,000 per month, payable in advance on the first day of each month."
        },
        {
            "question": "What is the amount of the security deposit?",
            "expected_answer": "The security deposit amount varies by agreement but typically ranges from $8,000 to $16,000, to be deposited upon execution of the lease."
        },
        {
            "question": "What are the permitted uses of the leased premises?",
            "expected_answer": "The tenant may use the premises solely for the specified purpose (e.g., retail store, restaurant, medical office, law office, or technology office) and for no other purpose without the landlord's prior written consent."
        },
        {
            "question": "Who is responsible for maintenance and repairs?",
            "expected_answer": "The tenant is responsible, at their own expense, for maintaining the premises in good condition and repair during the lease term. The landlord is responsible for structural repairs to the building."
        },
        {
            "question": "What insurance is the tenant required to maintain?",
            "expected_answer": "The tenant must maintain a policy of commercial general liability insurance for the premises with minimum liability limits of $1,000,000 to $3,000,000 per occurrence."
        }
    ],
    "Purchase Agreements": [
        {
            "question": "What is the purchase price for the property?",
            "expected_answer": "The purchase price varies by agreement but typically ranges from $850,000 to $2,000,000."
        },
        {
            "question": "What deposit is required when executing the agreement?",
            "expected_answer": "Upon execution of the agreement, the buyer must deposit an earnest money amount (typically $8,000 to $16,000) with the escrow agent."
        },
        {
            "question": "What happens if the buyer defaults on the agreement?",
            "expected_answer": "If the buyer defaults, the seller may, as its sole and exclusive remedy, terminate the agreement and retain the deposit as liquidated damages."
        },
        {
            "question": "What type of title will be conveyed to the buyer?",
            "expected_answer": "At closing, the seller shall convey good and marketable title to the property to the buyer by warranty deed, free and clear of all liens and encumbrances, except for those permitted by the agreement."
        },
        {
            "question": "What inspection rights does the buyer have?",
            "expected_answer": "The buyer has the right, at their expense, to conduct inspections of the property during the period from the agreement date until a specified inspection deadline."
        }
    ],
    "Shuttle Service Contract": [
        {
            "question": "What are the key services to be provided under the shuttle contract?",
            "expected_answer": "The contract involves the provision of shuttle transportation services, including vehicle operation, maintenance, and related services as specified in the agreement."
        },
        {
            "question": "What are the insurance requirements for the shuttle service?",
            "expected_answer": "The service provider must maintain various insurance coverages including commercial general liability, automobile liability, workers' compensation, and possibly umbrella liability policies with specified coverage limits."
        },
        {
            "question": "What are the termination provisions in the shuttle service contract?",
            "expected_answer": "The contract can typically be terminated for convenience with notice, for cause due to breach, for non-appropriation of funds, or for other specified reasons outlined in the termination section."
        },
        {
            "question": "What vehicles are required for the shuttle service?",
            "expected_answer": "The contract specifies requirements for shuttle vehicles including type, capacity, accessibility features, age limits, and compliance with relevant transportation regulations."
        },
        {
            "question": "What reporting requirements exist for the shuttle service provider?",
            "expected_answer": "The service provider must submit regular reports on ridership, service performance, maintenance, incidents, and other operational metrics as specified in the reporting and record-keeping sections."
        }
    ]
}

def create_evaluation_spreadsheet():
    """Create an Excel spreadsheet for evaluating document research queries."""
    # Create a dataframe to store all the evaluation data
    data = []
    
    # Add each document type and its questions
    for doc_type, questions in EVALUATION_QUESTIONS.items():
        files = ", ".join([os.path.basename(f) for f in DOCUMENT_TYPES[doc_type]])
        
        for i, question_data in enumerate(questions, 1):
            data.append({
                "Document Type": doc_type,
                "Files": files,
                "Question ID": f"{doc_type[:3]}-{i}",
                "Question": question_data["question"],
                "Expected Answer": question_data["expected_answer"],
                "Actual Answer": "",
                "Score (1-5)": "",
                "Notes": ""
            })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Write to Excel
    excel_file = "document_research_evaluation.xlsx"
    df.to_excel(excel_file, index=False)
    
    # Format the Excel file for better readability
    format_excel_file(excel_file)
    
    print(f"Created evaluation spreadsheet: {excel_file}")
    return excel_file

def format_excel_file(excel_file):
    """Format the Excel file for better readability."""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Document type fill
    doc_type_fills = {
        "Employment Contracts": PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid"),
        "Service Agreements": PatternFill(start_color="E6FFEF", end_color="E6FFEF", fill_type="solid"),
        "Non-Disclosure Agreements": PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"),
        "Lease Agreements": PatternFill(start_color="F2E6FF", end_color="F2E6FF", fill_type="solid"),
        "Purchase Agreements": PatternFill(start_color="FFECE6", end_color="FFECE6", fill_type="solid"),
        "Shuttle Service Contract": PatternFill(start_color="E6FFFF", end_color="E6FFFF", fill_type="solid")
    }
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Format data rows
    current_doc_type = None
    for row in range(2, ws.max_row + 1):
        doc_type = ws.cell(row=row, column=1).value
        
        # Apply fill based on document type
        fill = doc_type_fills.get(doc_type, None)
        if fill:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 20  # Document Type
    ws.column_dimensions[get_column_letter(2)].width = 30  # Files
    ws.column_dimensions[get_column_letter(3)].width = 10  # Question ID
    ws.column_dimensions[get_column_letter(4)].width = 40  # Question
    ws.column_dimensions[get_column_letter(5)].width = 50  # Expected Answer
    ws.column_dimensions[get_column_letter(6)].width = 50  # Actual Answer
    ws.column_dimensions[get_column_letter(7)].width = 10  # Score
    ws.column_dimensions[get_column_letter(8)].width = 30  # Notes
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the formatted workbook
    wb.save(excel_file)

if __name__ == "__main__":
    # Ensure pandas is installed
    try:
        import pandas as pd
        import openpyxl
    except ImportError:
        print("Please install required packages: pip install pandas openpyxl")
        sys.exit(1)
    
    create_evaluation_spreadsheet() 